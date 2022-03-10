import os
import sys
from collections import OrderedDict
from copy import deepcopy, copy

import openpyxl
from lxml import etree
from openpyxl.styles import *

person_in_charge = None
product_line = None

data_model_dic = {}
data_element_dic = {}
state_model_dic = OrderedDict()
publisher_dic = {}
line_cnt = 1
last_data_model_row_num = -1


def parse_data_element_to_json(data_element, ns, parent=""):
    name = data_element.attrib.get("name")
    if name is None:
        return
    if parent != "":
        full_name = parent + "." + name
    else:
        full_name = name
    data_dic = dict(data_element.attrib)
    data_dic["elementType"] = data_element.tag
    data_dic["fullName"] = full_name
    data_dic["parent"] = parent
    if data_dic["elementType"] in ["DataModel", "Block", "Flags", "Asn1Type", "Choice"]:
        data_dic["children"] = OrderedDict()

        for child in data_element.getchildren():
            tmp = parse_data_element_to_json(child, ns, full_name)
            # print(full_name)
            if tmp is not None:
                data_dic["children"][tmp["name"]] = tmp
    else:
        if data_element.find("Relation") is not None:
            relation_type = data_element.find("Relation").attrib["type"]
            relation_of = data_element.find("Relation").attrib["of"]
            data_dic["relationType"] = relation_type
            data_dic["relationOf"] = relation_of
    return data_dic


def parse_datamodel_in_statemodel(path, state_model_name, state_model_file):
    print("[+] Parsing datamodels in statemodel file %s, statemodel %s" % (state_model_file, state_model_name))
    global data_element_dic
    state_model_path = path + "/StateModels" + "/" + state_model_file
    data_model_path = path + "/DataModels"

    tree = etree.parse(state_model_path)
    data_model_nodes = tree.xpath("//StateModel[@name='%s']//DataModel" % state_model_name)

    if not data_model_nodes:
        print("找不到相应的DataModel")
    # print(len(data_model_nodes))
    ns = "default"
    data_model_dic[ns] = data_model_nodes
    data_element_dic[ns] = {}
    for data_element_node in data_model_nodes:
        if data_element_node.attrib.get("name") is None:
            print("状态模型中的数据模型也要有name属性")
            exit(0)
        tmp = parse_data_element_to_json(data_element=data_element_node, ns=ns, parent="")
        if tmp is not None:
            data_element_dic[ns][tmp["name"]] = tmp
    for include_node in tree.findall('//Include'):
        src_str = include_node.attrib.get("src")
        ns = include_node.attrib.get("ns")
        data_model_file = src_str[src_str.rfind("/") + 1:]
        data_model_dic[ns] = parse_datamodel(data_model_path + "/" + data_model_file)
        data_element_dic[ns] = {}
        for data_element_node in data_model_dic[ns]:
            tmp = parse_data_element_to_json(data_element=data_element_node, ns=ns, parent="")
            if tmp is not None:
                data_element_dic[ns][tmp["name"]] = tmp


def parse_datamodel(data_model_file):
    print("[+] Parsing datamodel file %s " % data_model_file)
    tree = etree.parse(data_model_file)
    for include_node in tree.findall('//Include'):
        ns = include_node.attrib.get("ns")
        src_str = include_node.attrib.get("src")
        included_data_model_file = src_str[src_str.rfind("/") + 1:]
        data_model_dic[ns] = parse_datamodel(os.path.dirname(data_model_file) + "\\" + included_data_model_file)
        data_element_dic[ns] = {}
        for data_element_node in data_model_dic[ns]:
            tmp = parse_data_element_to_json(data_element=data_element_node, ns=ns, parent="")
            data_element_dic[ns][tmp["name"]] = tmp

    return tree.xpath("//DataModel")


def merge_ref(ns, element):
    print("[+] Merge reference of %s:%s" % (ns, element.get("name")))
    src_ele = element
    if src_ele.get("children") is not None:
        for child in src_ele["children"]:
            merge_ref(ns, src_ele["children"][child])
    if src_ele.get("ref") is None:
        return
    ref_str = src_ele["ref"]
    if ":" not in ref_str:
        ref_ns = ns
        ref_name = ref_str
    else:
        ref_ns = ref_str.split(":")[0]
        ref_name = ref_str.split(":")[1]

    merge_ref(ref_ns, data_element_dic[ref_ns][ref_name])
    update_element_with_ref(src_ele, data_element_dic[ref_ns][ref_name])
    src_ele["ref"] = None


def get_child_with_xpath(xpath, root):
    paths = xpath.split(".")
    child_ele = root["children"].get(paths[0])
    if child_ele is None:
        return None
    for p in paths[1:]:
        children = child_ele.get("children")
        if children is None:
            return None
        child_ele = children.get(p)
        if child_ele is None:
            return None
    return child_ele


def update_element_with_ref(src_dic, ref_dic):
    print("[+] update %s element with %s datamodel" % (src_dic.get("name"), ref_dic.get("name")))
    ref_dic_copy = deepcopy(ref_dic)
    if src_dic.get("children") is not None:
        for child in src_dic["children"]:
            # print("child")
            # print(child)
            if child in ref_dic_copy["children"]:
                ref_dic_copy["children"][child] = deepcopy(src_dic["children"][child])
            else:
                ref_child_ele = get_child_with_xpath(child, ref_dic_copy)
                # print("ref_child_ele")
                # print(ref_child_ele)
                if ref_child_ele is not None:
                    ref_child_ele.update(src_dic["children"][child])
                    if src_dic["children"][child].get("children") is not None:
                        ref_child_ele["children"] = deepcopy(src_dic["children"][child]["children"])
                    ref_child_ele["name"] = src_dic["children"][child]["name"].split(".")[-1]
                else:
                    if "." in child:
                        ref_child_ele = get_child_with_xpath(".".join(child.split(".")[: -1]), ref_dic_copy)
                        child_name = child.split(".")[-1]
                        ref_child_ele["children"][child_name] = deepcopy(src_dic["children"][child])
                        ref_child_ele["children"][child_name]["name"] = child_name
                    else:
                        ref_dic_copy["children"][child] = src_dic["children"][child]

    src_name = src_dic["name"]
    src_full_name = src_dic["fullName"]
    src_type = src_dic["elementType"]

    src_dic.update(deepcopy(ref_dic_copy))
    src_dic["name"] = src_name
    src_dic["fullName"] = src_full_name
    if src_type == "Block" and src_dic["elementType"] == "DataModel":
        src_dic["elementType"] = "Block"
    update_full_names(src_dic)


def update_full_names(ele_dic):
    if ele_dic.get("children") is None:
        return
    full_name = ele_dic["fullName"]
    for child in ele_dic["children"]:
        ele_dic["children"][child]["fullName"] = full_name + "." + ele_dic["children"][child]["name"]
        ele_dic["children"][child]["parent"] = full_name
        update_full_names(ele_dic["children"][child])


def get_relative_name(parent_name, self_full_name):
    return self_full_name[len(parent_name) + 1:]


def add_data_model_to_sheet(worksheet, cur_data_dic):
    global line_cnt, last_data_model_row_num
    if cur_data_dic["elementType"] in ["Block", "DataModel", "Flags", "Asn1Type", "Choice"]:
        if cur_data_dic["elementType"] == "DataModel":
            worksheet["A" + str(line_cnt)] = cur_data_dic["name"]
            if last_data_model_row_num != -1:
                worksheet.merge_cells(start_row=last_data_model_row_num, start_column=1, end_row=line_cnt - 1,
                                      end_column=1)
                top_left_cell = worksheet["A" + str(last_data_model_row_num)]
                top_left_cell.alignment = Alignment(horizontal="left", vertical="top")
                top_left_cell.fill = PatternFill('solid', fgColor="EDDCDC")
                thin = Side(border_style="thin", color="D6D6D6")
                top_left_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            last_data_model_row_num = line_cnt
        if cur_data_dic.get("children") is not None:
            for child in cur_data_dic["children"]:
                add_data_model_to_sheet(worksheet, cur_data_dic["children"][child])
    else:
        worksheet["B" + str(line_cnt)] = cur_data_dic["name"]
        worksheet["C" + str(line_cnt)] = cur_data_dic["fullName"]
        worksheet["E" + str(line_cnt)] = cur_data_dic["elementType"]
        if cur_data_dic.get("length") is not None:
            worksheet["F" + str(line_cnt)] = int(cur_data_dic["length"]) * 8
        elif cur_data_dic.get("size") is not None:
            worksheet["F" + str(line_cnt)] = int(cur_data_dic["size"])

        if cur_data_dic.get("value") is not None:
            worksheet["H" + str(line_cnt)] = cur_data_dic["value"]

        if cur_data_dic.get("mutable") is not None:
            worksheet["I" + str(line_cnt)] = cur_data_dic.get("mutable")
        else:
            worksheet["I" + str(line_cnt)] = "true"

        if cur_data_dic.get("relationType") is not None:
            worksheet["J" + str(line_cnt)] = cur_data_dic["relationType"]
        if cur_data_dic.get("relationOf") is not None:
            worksheet["K" + str(line_cnt)] = cur_data_dic["relationOf"]
        line_cnt += 1


def add_state_model_to_sheet(worksheet):
    global line_cnt, state_model_dic, publisher_dic
    odd = True
    for state_model_name in state_model_dic:
        if len(state_model_dic[state_model_name]) == 0:
            continue
        worksheet["A" + str(line_cnt)] = state_model_name
        last_state_model_row_num = line_cnt
        for action in state_model_dic[state_model_name]:
            p_name = action["publisher"]
            if publisher_dic.get(p_name) is None:
                worksheet["E" + str(line_cnt)] = list(publisher_dic.values())[0]
            else:
                worksheet["E" + str(line_cnt)] = publisher_dic.get(p_name)
            if action["type"] == "input":
                worksheet["C" + str(line_cnt)] = action["data"]
                line_cnt += 1
                worksheet["C" + str(line_cnt)] = "<------------------------"
                line_cnt += 1
            elif action["type"] == "output":
                worksheet["C" + str(line_cnt)] = action["data"]
                line_cnt += 1
                worksheet["C" + str(line_cnt)] = "------------------------>"
                line_cnt += 1
            worksheet["C" + str(line_cnt)] = ""
            line_cnt += 1
            for col in range(5, 8):
                worksheet.merge_cells(start_row=line_cnt - 3, start_column=col, end_row=line_cnt - 1,
                                      end_column=col)

        worksheet.merge_cells(start_row=last_state_model_row_num, start_column=1, end_row=line_cnt - 1,
                              end_column=1)
        worksheet["A" + str(last_state_model_row_num)].font = Font(name='Calibri', bold=True)
        if odd:
            worksheet["A" + str(last_state_model_row_num)].fill = PatternFill('solid', fgColor="EDDCDC")
            odd = False
        else:
            worksheet["A" + str(last_state_model_row_num)].fill = PatternFill('solid', fgColor="F5EDF2")
            odd = True


def parse_state_model(path, state_model_name, state_model_file):
    global state_model_dic
    state_model_dic = OrderedDict()
    state_model_path = path + "/StateModels" + "/" + state_model_file
    root = etree.parse(state_model_path)
    states = root.xpath("//StateModel[@name='%s']/State" % state_model_name)
    for state in states:
        state_name = state.attrib.get("name")
        state_model_dic[state_name] = []
        for action in state.findall("Action"):
            type_ = action.attrib.get("type")
            if type_ == "input":
                action_dic = {"type": "input"}
                data_model = action.find("DataModel")
                action_dic["data"] = data_model.attrib.get("name")
                p_name = action.attrib.get("publisher")
                if p_name is None:
                    p_name = "default"
                action_dic["publisher"] = p_name
                state_model_dic[state_name].append(action_dic)
            elif type_ == "output":
                action_dic = {"type": "output"}
                data_model = action.find("DataModel")
                action_dic["data"] = data_model.attrib.get("name")
                state_model_dic[state_name].append(action_dic)
                p_name = action.attrib.get("publisher")
                if p_name is None:
                    p_name = "default"
                action_dic["publisher"] = p_name


def read_norm():
    work_book = openpyxl.load_workbook("norm.xlsx")
    work_sheet = work_book.active
    content = []
    for index, row in enumerate(work_sheet.rows):
        row_list = []
        for cell in row:
            row_list.append([cell.value, cell.font, cell.alignment, cell.border, cell.fill])
        content.append(row_list)
    return content


def parse_one_pit(pit_dir, pit_name, protocol_name):
    global line_cnt, last_data_model_row_num, data_element_dic, data_model_dic, state_model_dic, publisher_dic, \
        person_in_charge, product_line
    data_model_dic = {}
    data_element_dic = {}
    state_model_dic = OrderedDict()
    publisher_dic = OrderedDict()
    workbook = openpyxl.Workbook()
    line_cnt = 1
    last_data_model_row_num = -1
    pit_file = pit_dir + "/" + pit_name + ".xml"
    root = etree.parse(pit_file)
    if not root.xpath("/Secray"):
        print("xml根元素必须为Secray而不是Peach")
        exit(0)
    src = root.xpath("/Secray/Include")[0].attrib["src"]
    state_model_file = src[src.rfind("/") + 1:]
    state_model_name = root.xpath("//StateModel")[0].attrib["ref"].split(":")[1]
    publishers = root.xpath("//Publisher")
    for p in publishers:
        p_name = p.attrib.get("name")
        p_class = p.attrib.get("class")
        if p_name is None:
            p_name = "default"
        publisher_dic[p_name] = p_class
    # print(state_model_name)
    parse_datamodel_in_statemodel(pit_dir, state_model_name, state_model_file)
    # print(json.dumps(data_element_dic, indent=4))
    for name, dic in data_element_dic["default"].items():
        merge_ref("default", dic)

    # print(json.dumps(data_element_dic["default"], indent=4))

    base_info_worksheet = workbook.active
    base_info_worksheet.title = "测试套信息"

    base_info_worksheet["A1"] = "配置项名称"
    base_info_worksheet["B1"] = "配置项内容"
    base_info_worksheet["C1"] = "备注"

    base_info_worksheet["A2"] = "测试套名称"
    base_info_worksheet["B2"] = pit_name

    base_info_worksheet["A3"] = "测试套源代码存放位置"
    base_info_worksheet["A4"] = "协议名称"
    base_info_worksheet["B4"] = pit_name.split("_")[0]

    base_info_worksheet["A5"] = "协议版本"
    base_info_worksheet["A6"] = "测试套业务场景"
    base_info_worksheet["A7"] = "协议涉及官方文档"
    base_info_worksheet["A8"] = "开发维护责任人"
    if person_in_charge is not None:
        base_info_worksheet["B8"] = person_in_charge

    base_info_worksheet["A9"] = "所属产品线"
    if product_line is not None:
        base_info_worksheet["B9"] = product_line

    base_info_worksheet["A1"].font = Font(name='Calibri', bold=True)
    base_info_worksheet["A1"].fill = PatternFill('solid', fgColor="D7EAFA")
    base_info_worksheet["B1"].font = Font(name='Calibri', bold=True)
    base_info_worksheet["B1"].fill = PatternFill('solid', fgColor="BCDBF6")
    base_info_worksheet["C1"].font = Font(name='Calibri', bold=True)
    base_info_worksheet["C1"].fill = PatternFill('solid', fgColor="D7EAFA")
    base_info_worksheet.column_dimensions["A"].width = 50
    base_info_worksheet.column_dimensions["B"].width = 100
    base_info_worksheet.column_dimensions["C"].width = 60

    datamodel_worksheet = workbook.create_sheet(title="Data Model")
    # worksheet.title = pit_name
    datamodel_worksheet["A1"] = "Data Model（报文类型）"
    datamodel_worksheet["B1"] = "Field Name（字段名称）"
    datamodel_worksheet["C1"] = "Field Full Name（字段全称）"
    datamodel_worksheet["D1"] = "Description（字段说明）"
    datamodel_worksheet["E1"] = "Data Type（数据类型）"
    datamodel_worksheet["F1"] = "Data Length（数据长度/bit）"
    datamodel_worksheet["G1"] = "Legal Value（合法取值）"
    datamodel_worksheet["H1"] = "Default Value（默认值）"
    datamodel_worksheet["I1"] = "Mutable"
    datamodel_worksheet["J1"] = "Data Relation"
    datamodel_worksheet["K1"] = "Related Data Element"
    header_font = Font(name='Calibri', bold=True)
    back_color = PatternFill('solid', fgColor="D7EAFA")
    for cell in datamodel_worksheet[1]:
        cell.font = header_font
        cell.fill = back_color
    datamodel_worksheet.column_dimensions["A"].width = 30
    datamodel_worksheet.column_dimensions["B"].width = 30
    datamodel_worksheet.column_dimensions["C"].width = 60
    datamodel_worksheet.column_dimensions["D"].width = 30
    datamodel_worksheet.column_dimensions["E"].width = 25
    datamodel_worksheet.column_dimensions["F"].width = 25
    datamodel_worksheet.column_dimensions["G"].width = 25
    datamodel_worksheet.column_dimensions["H"].width = 30
    datamodel_worksheet.column_dimensions["I"].width = 20
    datamodel_worksheet.column_dimensions["J"].width = 20
    datamodel_worksheet.column_dimensions["K"].width = 20
    line_cnt += 1
    # print(data_element_dic)

    print("[+] Got %d datamodels for %s pit" % (len(data_element_dic["default"]), pit_name))
    for dm in data_element_dic["default"]:
        add_data_model_to_sheet(datamodel_worksheet, data_element_dic["default"][dm])
    datamodel_worksheet.merge_cells(start_row=last_data_model_row_num, start_column=1, end_row=line_cnt - 1,
                                    end_column=1)
    top_left_cell = datamodel_worksheet["A" + str(last_data_model_row_num)]
    top_left_cell.alignment = Alignment(horizontal="left", vertical="top")
    top_left_cell.fill = PatternFill('solid', fgColor="EDDCDC")
    thin = Side(border_style="thin", color="D6D6D6")
    top_left_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    statemodel_worksheet = workbook.create_sheet("State Model")
    statemodel_worksheet["A1"] = "State Name"
    statemodel_worksheet.merge_cells(start_row=1, start_column=1, end_row=2,
                                     end_column=1)
    statemodel_worksheet.merge_cells(start_row=1, start_column=3, end_row=2,
                                     end_column=3)
    statemodel_worksheet.merge_cells(start_row=1, start_column=5, end_row=2,
                                     end_column=5)
    statemodel_worksheet.merge_cells(start_row=1, start_column=6, end_row=2,
                                     end_column=6)
    statemodel_worksheet.merge_cells(start_row=1, start_column=7, end_row=2,
                                     end_column=7)
    statemodel_worksheet["B1"] = "Endpoint"
    statemodel_worksheet["C1"] = "Packet"
    statemodel_worksheet["D1"] = "Endpoint"
    statemodel_worksheet["B2"] = "SecRAY"
    statemodel_worksheet["D2"] = "Target"
    statemodel_worksheet["E1"] = "Publisher"
    statemodel_worksheet["F1"] = "报文处理"
    statemodel_worksheet["G1"] = "报文交互说明"
    statemodel_worksheet["A1"].font = header_font
    statemodel_worksheet["A1"].fill = back_color

    thin = Side(border_style="thin", color="FFFFFF")
    for cell in statemodel_worksheet[1]:
        cell.font = header_font
        cell.fill = back_color
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for cell in statemodel_worksheet[2]:
        cell.font = header_font
        cell.fill = back_color
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    line_cnt = 3

    parse_state_model(pit_dir, state_model_name, state_model_file)
    add_state_model_to_sheet(statemodel_worksheet)
    for row in statemodel_worksheet:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in statemodel_worksheet["B3:B" + str(line_cnt - 1)]:
        cell[0].fill = PatternFill('solid', fgColor="DEF7E2")
    for cell in statemodel_worksheet["C3:C" + str(line_cnt - 1)]:
        cell[0].fill = PatternFill('solid', fgColor="FFFFFF")
    for cell in statemodel_worksheet["D3:D" + str(line_cnt - 1)]:
        cell[0].fill = PatternFill('solid', fgColor="DEF7E2")
    statemodel_worksheet.column_dimensions["A"].width = 20
    statemodel_worksheet.column_dimensions["B"].width = 20
    statemodel_worksheet.column_dimensions["C"].width = 30
    statemodel_worksheet.column_dimensions["D"].width = 20
    statemodel_worksheet.column_dimensions["E"].width = 20
    statemodel_worksheet.column_dimensions["F"].width = 50
    statemodel_worksheet.column_dimensions["G"].width = 50

    new_worksheet = workbook.create_sheet(title="测试套开发规范")
    norm_workbook = openpyxl.load_workbook("norm.xlsx")
    norm_worksheet = norm_workbook.active
    for row_num, row in enumerate(norm_worksheet.rows):
        new_worksheet.row_dimensions[row_num].height = norm_worksheet.row_dimensions[row_num].height
        for col_num, cell in enumerate(row):
            new_cell = new_worksheet.cell(row=row_num + 1, column=col_num + 1, value=cell.value)
            new_cell.font = copy(cell.font)
            new_cell.alignment = copy(cell.alignment)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
    new_worksheet.merged_cells = copy(norm_worksheet.merged_cells)
    new_worksheet.column_dimensions["A"].width = norm_worksheet.column_dimensions["A"].width
    new_worksheet.column_dimensions["B"].width = norm_worksheet.column_dimensions["B"].width
    new_worksheet.column_dimensions["C"].width = norm_worksheet.column_dimensions["C"].width
    new_worksheet.column_dimensions["D"].width = norm_worksheet.column_dimensions["D"].width
    new_worksheet.column_dimensions["E"].width = norm_worksheet.column_dimensions["E"].width
    workbook.save(protocol_name + "/" + pit_name + "_测试套设计文档.xlsx")


def main():
    global person_in_charge, product_line
    protocol_dir = sys.argv[1]
    if len(sys.argv) > 2:
        person_in_charge = sys.argv[2]
        product_line = sys.argv[3]
    protocol_name = os.path.basename(protocol_dir)
    if not os.path.exists(protocol_name):
        os.mkdir(protocol_name)
    # print(norm_content)
    for file in os.listdir(protocol_dir):
        if file.endswith(".xml") and not file.endswith(".config.xml"):
            pit_name = os.path.basename(file)[: -4]
            parse_one_pit(pit_dir=protocol_dir, pit_name=pit_name, protocol_name=protocol_name)


def func(pit_dir):
    protocol_name = os.path.basename(pit_dir)
    if not os.path.exists(pit_dir + "/DataModels") or not os.path.exists(pit_dir + "/StateModels"):
        print("必须按照规范目录结构存放")
        exit(0)
    for file in os.listdir(pit_dir):
        if file.endswith(".xml") and not file.endswith(".config.xml"):
            pit_name = os.path.basename(file)[: -4]
            parse_one_pit(pit_dir=pit_dir, pit_name=pit_name, protocol_name=protocol_name)


if __name__ == "__main__":
    main()
